"""Parse an AdImpact "Spending Chart" export into a tidy long-format dataset.

The export is a collapsed pivot table: Party -> AdvertiserType -> Advertiser ->
Market -> MediaType -> Station, with a Grand Total column followed by one
column per week. Blank cells mean "same as the row above" and AdImpact
inserts a "<LEVEL> TOTAL" subtotal row after every group. We forward-fill the
hierarchy, drop the subtotal rows (they're redundant with the leaf rows), and
re-derive a continuous weekly axis (AdImpact drops weeks that are $0 across
every advertiser instead of keeping a fixed weekly grid).
"""
import re
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import openpyxl

# Media-week convention (matches GPS Impact's linear/compbot reports):
# Tuesday is the first day of the reporting week, Tuesday -> Monday, and
# Eastern time governs the rollover regardless of where this runs.
MEDIA_WEEK_TZ = ZoneInfo("America/New_York")


def _tuesday_on_or_before(d):
    """The Tuesday on or before date/datetime `d` (media-week start)."""
    days_back = (d.isoweekday() - 2) % 7  # Mon=1 ... Sun=7; walk back to Tue
    return d - timedelta(days=days_back)


def current_media_week_iso(now=None):
    """ISO date (YYYY-MM-DD) of the Tuesday on or before `now`.

    `now` is for tests; production uses the current America/New_York wall
    clock. This is "this week" as GPS Impact's staff mean it, not whichever
    week happens to be the rightmost column in a given export — those are
    only the same thing when the export is fully caught up to today.
    """
    if now is None:
        now = datetime.now(MEDIA_WEEK_TZ)
    return _tuesday_on_or_before(now).strftime("%Y-%m-%d")

HIERARCHY_COLS = ("party", "atype", "advertiser", "market", "mediatype", "station")


def _num(v):
    return float(v) if v not in (None, "") else 0.0


def load_spending_export(path, sheet_name=None):
    """Read the raw export and return (week_cols, week_starts, leaf_rows, meta).

    leaf_rows is a list of dicts with keys: party, atype, advertiser, market,
    mediatype, station, grand_total, weeks (tuple aligned to week_cols).
    meta carries whatever of race / media_types / export_date the sheet's
    header block declared (any of them may be None if the export omits it).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    header_row_idx = None
    meta = {"race": None, "media_types": None, "export_date": None}
    for row in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row)):
        v = row[0].value
        if v == "Party":
            header_row_idx = row[0].row
            break
        if isinstance(v, str):
            m = re.match(r"Race:\s*(.+)", v)
            if m:
                meta["race"] = m.group(1).strip()
            m = re.match(r"Media Types?:\s*(.+)", v)
            if m:
                meta["media_types"] = m.group(1).strip()
            m = re.match(r"Export Date:\s*(.+)", v)
            if m:
                meta["export_date"] = m.group(1).strip()
    if header_row_idx is None:
        raise ValueError(
            "Could not find the 'Party' header row — is this an AdImpact "
            "Spending Chart export grouped by Party/AdvertiserType/Advertiser/"
            "Market/MediaType/Station?"
        )

    header = [c.value for c in ws[header_row_idx]]
    week_cols = header[7:]
    week_starts = []
    for w in week_cols:
        m = re.match(r"(\d{2}/\d{2}/\d{4})", str(w))
        if not m:
            raise ValueError(f"Could not parse a week start date from column header {w!r}")
        week_starts.append(m.group(1))

    data_start = header_row_idx + 2  # skip the header row + the "Spend" subheader row
    rows = list(ws.iter_rows(min_row=data_start, max_row=ws.max_row, values_only=True))

    leaf_rows = []
    cur = [None] * 6
    for r in rows:
        if all(v is None for v in r[:6]):
            continue
        if isinstance(r[0], str) and ("adimpact" in r[0].lower() or "confidential" in r[0].lower()):
            break
        vals = list(r[:6])
        for i in range(6):
            if vals[i] not in (None, ""):
                cur[i] = vals[i]
        station = vals[5]
        is_total = any(isinstance(v, str) and "TOTAL" in v.upper() for v in vals)
        if is_total or station is None:
            continue
        leaf_rows.append(
            {
                "party": cur[0],
                "atype": cur[1],
                "advertiser": cur[2],
                "market": cur[3],
                "mediatype": cur[4],
                "station": cur[5],
                "grand_total": r[6],
                "weeks": tuple(_num(x) for x in r[7 : 7 + len(week_cols)]),
            }
        )

    return week_cols, week_starts, leaf_rows, meta


def build_continuous_week_axis(week_starts):
    """Fill any $0-everywhere weeks AdImpact dropped so the axis stays weekly.

    Returns (full_week_labels, full_week_iso, index_map) where index_map[i]
    is the position of original week i within the continuous axis.
    """
    dates = [datetime.strptime(w, "%m/%d/%Y") for w in week_starts]
    full = []
    d = dates[0]
    while d <= dates[-1]:
        full.append(d)
        d += timedelta(days=7)
    pos = {d: i for i, d in enumerate(full)}
    index_map = [pos[d] for d in dates]
    labels = [d.strftime("%-m/%-d/%y") for d in full]
    iso = [d.strftime("%Y-%m-%d") for d in full]
    return labels, iso, index_map


def aggregate(leaf_rows, index_map, n_weeks, top_n=6):
    """Roll leaf rows up into platform-week, advertiser-week and top-N series.

    Returns a dict ready to be handed to report.colors.assign_colors() and
    then serialized as the report's DATA payload.
    """

    def remap(values):
        out = [0.0] * n_weeks
        for i, v in enumerate(values):
            out[index_map[i]] += v
        return out

    platform_week = {}
    advertiser_week = {}
    advertiser_meta = {}
    advertiser_platform = {}

    for r in leaf_rows:
        adv, st, party, atype = r["advertiser"], r["station"], r["party"], r["atype"]
        advertiser_meta[adv] = {"party": party, "atype": atype}
        vals = remap(r["weeks"])
        platform_week.setdefault(st, [0.0] * n_weeks)
        advertiser_week.setdefault(adv, [0.0] * n_weeks)
        advertiser_platform.setdefault(adv, {})
        for i in range(n_weeks):
            platform_week[st][i] += vals[i]
            advertiser_week[adv][i] += vals[i]
        advertiser_platform[adv][st] = advertiser_platform[adv].get(st, 0.0) + sum(vals)

    advertiser_total = {a: sum(w) for a, w in advertiser_week.items()}
    grand_total = sum(advertiser_total.values())
    ranked = sorted(advertiser_total.items(), key=lambda x: -x[1])

    r2 = lambda x: round(x, 2)

    platform_series = [
        {"key": p, "total": r2(sum(v)), "values": [r2(x) for x in v]}
        for p, v in sorted(platform_week.items(), key=lambda kv: -sum(kv[1]))
    ]

    advertiser_rows = []
    for a, total in ranked:
        m = advertiser_meta[a]
        advertiser_rows.append(
            {
                "name": a,
                "party": m["party"],
                "atype": m["atype"],
                "total": r2(total),
                "share": r2(total / grand_total * 100) if grand_total else 0.0,
                "weekly": [r2(v) for v in advertiser_week[a]],
                "platform": {k: r2(v) for k, v in advertiser_platform[a].items()},
            }
        )

    top_advs = [a for a, _ in ranked[:top_n]]
    other_advs = [a for a, _ in ranked[top_n:]]
    top_series = []
    for a in top_advs:
        top_series.append(
            {
                "name": a,
                "party": advertiser_meta[a]["party"],
                "atype": advertiser_meta[a]["atype"],
                "total": r2(advertiser_total[a]),
                "values": [r2(v) for v in advertiser_week[a]],
            }
        )
    if other_advs:
        others_weekly = [r2(sum(advertiser_week[a][i] for a in other_advs)) for i in range(n_weeks)]
        top_series.append(
            {
                "name": "All Other Advertisers",
                "party": "Other",
                "atype": "Mixed",
                "total": r2(sum(others_weekly)),
                "values": others_weekly,
            }
        )

    return {
        "grand_total": r2(grand_total),
        "platform_series": platform_series,
        "advertiser_rows": advertiser_rows,
        "top_series": top_series,
        "other_count": len(other_advs),
    }


def load_creative_export(path, sheet_name=None):
    """Read an AdImpact "Topline Creatives" export.

    Returns a list of dicts, one per creative: start, end (datetime),
    title, advertiser, platform ("CTV" or "Digital" — whichever of the two
    spend columns is nonzero; a creative is never both in this export),
    tone, total_spend, url (AdImpact's public ad-preview link, or None if
    the export doesn't have that column). Broadcast/Cable columns are read
    but discarded — this export is meant to be pulled already filtered to
    Media Type = CTV, Digital, so those columns are $0 anyway, not just
    irrelevant.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row)):
        if row[0].value == "Start Date":
            header_row_idx = row[0].row
            break
    if header_row_idx is None:
        raise ValueError(
            "Could not find the 'Start Date' header row — is this an AdImpact "
            "Topline Creatives export?"
        )
    header = [c.value for c in ws[header_row_idx]]
    col = {name: i for i, name in enumerate(header)}

    rows = []
    for r in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=True):
        start, end, advertiser = r[col["Start Date"]], r[col["End Date"]], r[col["Advertiser"]]
        if not isinstance(start, datetime) or not isinstance(advertiser, str):
            continue
        ctv_spend = _num(r[col["CTV Spend"]])
        digital_spend = _num(r[col["Digital Spend"]])
        rows.append(
            {
                "start": start,
                "end": end if isinstance(end, datetime) else start,
                "title": r[col["Title"]] or "(untitled)",
                "advertiser": advertiser,
                "platform": "CTV" if ctv_spend >= digital_spend else "Digital",
                "tone": r[col["Tone"]],
                "total_spend": _num(r[col["Total Spend"]]),
                "url": r[col["Public URL"]] if "Public URL" in col else None,
            }
        )
    return rows


def build_creative_week_axis(creative_rows):
    """Continuous Tuesday-start weekly axis spanning every creative's flight.

    Returns (week_iso, week_labels) — week_iso for lookups, week_labels
    ("M/D", no year — these head narrow Gantt-bar columns with no room
    for one) both aligned to the same Tuesday->Monday media-week
    convention as the spending report.
    """
    starts = [_tuesday_on_or_before(r["start"]) for r in creative_rows]
    ends = [_tuesday_on_or_before(r["end"]) for r in creative_rows]
    first, last = min(starts), max(ends)
    weeks = []
    d = first
    while d <= last:
        weeks.append(d)
        d += timedelta(days=7)
    week_iso = [d.strftime("%Y-%m-%d") for d in weeks]
    week_labels = [d.strftime("%-m/%-d") for d in weeks]
    return week_iso, week_labels


def creative_active_weeks(creative, week_iso):
    """Indices into `week_iso` where `creative` overlaps that media week."""
    start, end = creative["start"], creative["end"]
    active = []
    for i, iso in enumerate(week_iso):
        week_start = datetime.strptime(iso, "%Y-%m-%d")
        week_end = week_start + timedelta(days=6)
        if start <= week_end and end >= week_start:
            active.append(i)
    return active
