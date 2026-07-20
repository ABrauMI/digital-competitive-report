"""Write the flat, spreadsheet-style competitive report GPS Impact already
produces for linear TV — same hierarchy, same shade ramp by rollup level,
same merged-cell nesting — reapplied to digital spend, in GPS Impact's own
2026 brand colors and with the brand logo embedded in the header band.

Layout (per the linear "Competitive TV Report" template this mirrors):
  CANDIDATE / COMMITTEE -> MARKET -> TYPE -> STATION/PLATFORM, with a Total
  Spend column and one column per week. Subtotal rows roll up Market+Type,
  then the advertiser, then the party, then a grand total.

Digital has no GRPs/CPP equivalent without a separate impressions export, so
those columns are simply omitted for now — spend only, matching how far the
underlying AdImpact export goes.
"""
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import parse

# GPS Impact Brand Guidelines 2026
BRAND_NAVY = "323b51"
BRAND_BLUE = "3d6a91"
BRAND_RED = "de5e4e"
BRAND_PALE_BLUE = "bed7d5"

NAVY = BRAND_NAVY
HEADER_ACCENT = BRAND_RED
GRAND_FILL = BRAND_NAVY
TEXT_DARK = BRAND_NAVY
BORDER_LIGHT = "E5E7EB"
BORDER_MED = "CCCCCC"
FOOTER_GRAY = "6B7280"

# Each advertiser's whole block (leaf -> market/type subtotal -> type total
# -> advertiser total) is shaded in a light-to-dark tint ramp of its party's
# color instead of one blue ramp for everyone — a shade of GOP red for
# Republican advertisers, brand blue for Democrats, neutral gray for
# anything else (Independent, Nonpartisan, ...). Each ramp is four tints
# (10%/22%/30%/38% toward white) of the party color, landing on the full
# color itself for the party-total row. See README.md#color-choices.
PARTY_RAMPS = {
    "Republican": {
        "leaf": "fcefed", "subtotal": "f8dcd8", "type_total": "f5cfca", "adv_total": "f2c2bc",
        "party_total": BRAND_RED,
    },
    "Democrat": {
        "leaf": "ecf0f4", "subtotal": "d4dee7", "type_total": "c5d2de", "adv_total": "b5c6d5",
        "party_total": BRAND_BLUE,
    },
}
DEFAULT_RAMP = {
    "leaf": "f3f4f6", "subtotal": "e5e6ea", "type_total": "dcdde2", "adv_total": "d3d4db",
    "party_total": "8a8fa0",
}


def _ramp_for(party):
    return PARTY_RAMPS.get(party, DEFAULT_RAMP)

LEAF_CURRENCY = '$#,##0;-$#,##0;""'
TOTAL_CURRENCY = "$#,##0"

FONT_NAME = "Calibri"

FIRST_DATA_ROW = 4  # 1=title, 2=blank, 3=header

ASSETS_DIR = Path(__file__).resolve().parent.parent / "assets"
LOGO_FILE = ASSETS_DIR / "logos" / "GPSImpact_White_Horizontal_2026.png"


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _thin_border(color):
    side = Side(style="thin", color=color)
    return Border(top=side, bottom=side)


def _medium_border(color):
    side = Side(style="medium", color=color)
    return Border(top=side, bottom=side)


def _sum_lists(lists, n_weeks):
    out = [0.0] * n_weeks
    for lst in lists:
        for i in range(n_weeks):
            out[i] += lst[i]
    return out


def _sum_pairs(pairs):
    pairs = list(pairs)
    return sum(p[0] for p in pairs), sum(p[1] for p in pairs)


def _filter_this_week(tree, this_idx, prior_idx):
    """Collapse the tree to (this_week, prior_week) leaf pairs, dropping any
    branch with nothing spent this week — "what's happening" means active
    now, not everything that's ever run."""
    out = {}
    for party, advertisers in tree.items():
        adv_out = {}
        for adv, markets in advertisers.items():
            mk_out = {}
            for market, types in markets.items():
                tp_out = {}
                for mtype, stations in types.items():
                    st_out = {}
                    for station, weekly in stations.items():
                        this_val = weekly[this_idx]
                        prior_val = weekly[prior_idx] if prior_idx is not None and prior_idx >= 0 else 0.0
                        if this_val > 0:
                            st_out[station] = (this_val, prior_val)
                    if st_out:
                        tp_out[mtype] = st_out
                if tp_out:
                    mk_out[market] = tp_out
            if mk_out:
                adv_out[adv] = mk_out
        if adv_out:
            out[party] = adv_out
    return out


def build_hierarchy(leaf_rows, index_map, n_weeks):
    """party -> advertiser -> market -> mediatype -> station -> weekly[]."""

    def remap(values):
        out = [0.0] * n_weeks
        for i, v in enumerate(values):
            out[index_map[i]] += v
        return out

    tree = {}
    for r in leaf_rows:
        party_node = tree.setdefault(r["party"], {})
        adv_node = party_node.setdefault(r["advertiser"], {})
        market_node = adv_node.setdefault(r["market"], {})
        type_node = market_node.setdefault(r["mediatype"], {})
        arr = type_node.setdefault(r["station"], [0.0] * n_weeks)
        vals = remap(r["weeks"])
        for i in range(n_weeks):
            arr[i] += vals[i]
    return tree


def build_hierarchy_no_market(leaf_rows, n_weeks):
    """party -> advertiser -> mediatype -> station -> weekly[].

    Same shape as `build_hierarchy` minus the market level, for sources
    like AdHawk that don't carry a market/DMA column at all. Leaf rows'
    `weekly` arrays are expected already aligned to a single n_weeks axis
    (no index_map remap needed — AdHawk doesn't drop $0 weeks the way
    AdImpact does).
    """
    tree = {}
    for r in leaf_rows:
        party_node = tree.setdefault(r["party"], {})
        adv_node = party_node.setdefault(r["advertiser"], {})
        type_node = adv_node.setdefault(r["mediatype"], {})
        arr = type_node.setdefault(r["station"], [0.0] * n_weeks)
        for i in range(n_weeks):
            arr[i] += r["weekly"][i]
    return tree


def _filter_this_week_no_market(tree, this_idx, prior_idx):
    """`_filter_this_week`, minus the market level."""
    out = {}
    for party, advertisers in tree.items():
        adv_out = {}
        for adv, types in advertisers.items():
            tp_out = {}
            for mtype, stations in types.items():
                st_out = {}
                for station, weekly in stations.items():
                    this_val = weekly[this_idx]
                    prior_val = weekly[prior_idx] if prior_idx is not None and prior_idx >= 0 else 0.0
                    if this_val > 0:
                        st_out[station] = (this_val, prior_val)
                if st_out:
                    tp_out[mtype] = st_out
            if tp_out:
                adv_out[adv] = tp_out
        if adv_out:
            out[party] = adv_out
    return out


MAIN_SHEET_HEADERS = ["CANDIDATE / COMMITTEE", "MARKET", "TYPE", "STATION / PLATFORM", "TOTAL SPEND"]
MAIN_SHEET_WIDTHS = {"A": 28, "B": 24, "C": 10, "D": 30, "E": 14}

# AdHawk has no Market/DMA column at all — it's digital-only spend, and
# digital doesn't target by broadcast DMA the way linear TV does — so its
# sheet is one column shallower than AdImpact's.
ADHAWK_MAIN_SHEET_HEADERS = ["CANDIDATE / COMMITTEE", "TYPE", "PLATFORM", "TOTAL SPEND"]
ADHAWK_MAIN_SHEET_WIDTHS = {"A": 32, "B": 10, "C": 24, "D": 14}


class _Writer:
    def __init__(self, ws, n_weeks, week_labels, title, headers=None, col_widths=None):
        self.ws = ws
        self.n_weeks = n_weeks
        self.week_labels = week_labels
        self.row = 1
        self.ramp = DEFAULT_RAMP
        headers = headers or MAIN_SHEET_HEADERS
        col_widths = col_widths or MAIN_SHEET_WIDTHS
        self.frozen_cols = len(headers)  # last fixed column is always TOTAL SPEND
        self._write_scaffold(title, headers, col_widths)

    def set_party(self, party):
        """Every write_* call after this uses `party`'s color ramp."""
        self.ramp = _ramp_for(party)

    def _write_scaffold(self, title, headers, col_widths):
        ws = self.ws
        ws.sheet_view.showGridLines = False
        frozen_cols = self.frozen_cols
        last_col = frozen_cols + self.n_weeks

        for col in range(1, last_col + 1):
            ws.cell(1, col).fill = _fill(NAVY)

        # Two merges, not one: a merged cell can't straddle a freeze-pane
        # split without rendering oddly once you scroll, so the title stays
        # inside the frozen columns and the rest of the navy band is a
        # second, separate merge entirely in the scrollable region.
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=frozen_cols)
        c = ws.cell(1, 1, title)
        c.font = Font(name=FONT_NAME, bold=True, size=14, color="FFFFFF")
        c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True, indent=1)
        if last_col > frozen_cols:
            ws.merge_cells(start_row=1, start_column=frozen_cols + 1, end_row=1, end_column=last_col)
        ws.row_dimensions[1].height = 49.5

        if LOGO_FILE.exists():
            logo = XLImage(str(LOGO_FILE))
            aspect = logo.width / logo.height
            logo.height = 34
            logo.width = 34 * aspect
            ws.add_image(logo, "A1")

        for col, text in enumerate(headers + self.week_labels, start=1):
            c = ws.cell(3, col, text)
            c.font = Font(name=FONT_NAME, bold=True, size=9, color="FFFFFF")
            c.fill = _fill(NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _medium_border(HEADER_ACCENT)
        ws.row_dimensions[3].height = 31.5

        for col_letter, w in col_widths.items():
            ws.column_dimensions[col_letter].width = w
        for i in range(self.n_weeks):
            ws.column_dimensions[get_column_letter(frozen_cols + 1 + i)].width = 13

        ws.freeze_panes = f"{get_column_letter(frozen_cols + 1)}{FIRST_DATA_ROW}"
        self.row = FIRST_DATA_ROW

    def _write_row(self, total_value, weekly_values, fill, bold, text_color, border, font_size=9,
                    currency=LEAF_CURRENCY, label=None, label_col=None):
        ws = self.ws
        r = self.row
        total_col = self.frozen_cols
        font = Font(name=FONT_NAME, bold=bold, size=font_size, color=text_color)
        last_col = total_col + self.n_weeks
        for col in range(1, last_col + 1):
            c = ws.cell(r, col)
            c.fill = _fill(fill)
            c.border = border
            c.font = font
        if label is not None:
            ws.cell(r, label_col if label_col is not None else total_col - 1, label)
        e = ws.cell(r, total_col, total_value)
        e.number_format = currency
        for i, v in enumerate(weekly_values):
            cc = ws.cell(r, total_col + 1 + i, v)
            cc.number_format = currency
        self.row += 1
        return r

    def write_leaf(self, station, total, weekly):
        r = self.row
        ws = self.ws
        leaf_fill = self.ramp["leaf"]
        label_col = self.frozen_cols - 1
        total_col = self.frozen_cols
        ws.cell(r, label_col, station).font = Font(name=FONT_NAME, size=9, color=TEXT_DARK)
        for col in range(1, total_col + 1):
            ws.cell(r, col).fill = _fill(leaf_fill)
            ws.cell(r, col).border = _thin_border(BORDER_LIGHT)
        ws.cell(r, total_col, total).number_format = LEAF_CURRENCY
        ws.cell(r, total_col).font = Font(name=FONT_NAME, size=9, color=TEXT_DARK)
        for i, v in enumerate(weekly):
            cc = ws.cell(r, total_col + 1 + i, v if v else None)
            cc.number_format = LEAF_CURRENCY
            cc.font = Font(name=FONT_NAME, size=9, color=TEXT_DARK)
            cc.fill = _fill(leaf_fill)
            cc.border = _thin_border(BORDER_LIGHT)
        self.row += 1
        return r

    def write_subtotal(self, label, total, weekly):
        return self._write_row(total, weekly, self.ramp["subtotal"], True, TEXT_DARK, _thin_border(BORDER_MED),
                                currency=TOTAL_CURRENCY, label=label)

    def write_type_total(self, label, total, weekly):
        return self._write_row(total, weekly, self.ramp["type_total"], True, TEXT_DARK, _thin_border(BORDER_MED),
                                currency=TOTAL_CURRENCY, label=label)

    def write_advertiser_total(self, label, total, weekly):
        return self._write_row(total, weekly, self.ramp["adv_total"], True, TEXT_DARK, _medium_border(BORDER_MED),
                                currency=TOTAL_CURRENCY, label=label)

    def write_party_total(self, label, total, weekly):
        return self._write_row(
            total, weekly, self.ramp["party_total"], True, "FFFFFF", _medium_border("FFFFFF"), font_size=10,
            currency=TOTAL_CURRENCY, label=label, label_col=1,
        )

    def write_grand_total(self, label, total, weekly):
        return self._write_row(
            total, weekly, GRAND_FILL, True, "FFFFFF", _medium_border("FFFFFF"), font_size=10,
            currency=TOTAL_CURRENCY, label=label, label_col=1,
        )

    def blank_row(self):
        self.row += 1

    def write_footer(self):
        last_col = self.frozen_cols + self.n_weeks
        self.blank_row()
        r = self.row
        self.ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        c = self.ws.cell(r, 1, "Report prepared by GPS Impact  |  Confidential")
        c.font = Font(name=FONT_NAME, italic=True, size=8, color=FOOTER_GRAY)
        self.ws.row_dimensions[r].height = 18
        self.row += 1

    def write_group_label(self, col, row, text):
        c = self.ws.cell(row, col, text)
        c.font = Font(name=FONT_NAME, bold=True, size=9, color=TEXT_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def merge_col(self, col, start_row, end_row):
        if end_row > start_row:
            self.ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
            top = self.ws.cell(start_row, col)
            top.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def merge_label_row(self, row, start_col=1, end_col=None):
        if end_col is None:
            end_col = self.frozen_cols - 1
        self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)


def _write_main_sheet(wb, tree, week_labels, n_weeks, title):
    ws = wb.active
    ws.title = "Competitive Digital Report"
    w = _Writer(ws, n_weeks, week_labels, title)

    parties = sorted(tree.items(), key=lambda kv: -sum(_sum_lists(
        [s for adv in kv[1].values() for mk in adv.values() for tp in mk.values() for s in tp.values()], n_weeks
    )))

    for party, advertisers in parties:
        w.set_party(party)
        adv_totals = {
            adv: _sum_lists([s for mk in markets.values() for tp in mk.values() for s in tp.values()], n_weeks)
            for adv, markets in advertisers.items()
        }
        ranked_advs = sorted(advertisers.items(), key=lambda kv: -sum(adv_totals[kv[0]]))

        party_weekly = _sum_lists(adv_totals.values(), n_weeks)

        for adv, markets in ranked_advs:
            adv_start = w.row
            w.write_group_label(1, adv_start, adv)

            # Group by media type first so every CTV market is together,
            # every Digital market is together, and each group gets its own
            # rolled-up total for the advertiser.
            by_type = {}
            for market, types in markets.items():
                for mtype, stations in types.items():
                    by_type.setdefault(mtype, {})[market] = stations
            ranked_types = sorted(
                by_type.items(),
                key=lambda kv: -sum(sum(weekly) for stations in kv[1].values() for weekly in stations.values()),
            )

            for mtype, market_map in ranked_types:
                for market, stations in market_map.items():
                    market_start = w.row
                    w.write_group_label(2, market_start, market)
                    type_start = w.row
                    w.write_group_label(3, type_start, mtype)
                    type_weekly = _sum_lists(stations.values(), n_weeks)

                    if mtype.strip().upper() == "CTV":
                        # All CTV platforms (In-App, Device, Streaming, ...)
                        # roll into one combined line per market rather than
                        # breaking out by individual platform. It's a plain
                        # data row now, not a rollup of visible rows above
                        # it, so it gets leaf styling/label, not subtotal.
                        w.write_leaf(f"{market} - {mtype}", sum(type_weekly), type_weekly)
                    else:
                        ranked_stations = sorted(stations.items(), key=lambda kv: -sum(kv[1]))
                        for station, weekly in ranked_stations:
                            w.write_leaf(station, sum(weekly), weekly)
                        type_end = w.row - 1
                        w.merge_col(3, type_start, type_end)  # TYPE column
                        w.write_subtotal(f"{market} - {mtype} Total", sum(type_weekly), type_weekly)
                    market_end = w.row - 1
                    w.merge_col(2, market_start, market_end)  # MARKET column

                type_total_weekly = _sum_lists(
                    [s for stations in market_map.values() for s in stations.values()], n_weeks
                )
                w.write_type_total(f"{mtype.upper()} TOTAL", sum(type_total_weekly), type_total_weekly)

            w.write_advertiser_total(f"{adv} Total", sum(adv_totals[adv]), adv_totals[adv])
            adv_end = w.row - 1
            w.merge_col(1, adv_start, adv_end)  # CANDIDATE / COMMITTEE column
            w.blank_row()

        w.write_party_total(f"{party.upper()} PARTY TOTAL", sum(party_weekly), party_weekly)
        w.merge_label_row(w.row - 1)

    grand_weekly = _sum_lists(
        [_sum_lists(
            [s for adv in advertisers.values() for mk in adv.values() for tp in mk.values() for s in tp.values()],
            n_weeks,
        ) for _, advertisers in parties],
        n_weeks,
    )
    w.write_grand_total("GRAND TOTAL", sum(grand_weekly), grand_weekly)
    w.merge_label_row(w.row - 1)
    w.write_footer()
    return ws


def _write_main_sheet_adhawk(wb, tree, week_labels, n_weeks, title):
    """Same layout as `_write_main_sheet`, minus the Market column — AdHawk
    has no market/DMA data at all, so the hierarchy goes straight from
    advertiser to CTV/Digital to platform."""
    ws = wb.active
    ws.title = "Competitive Digital Report"
    w = _Writer(ws, n_weeks, week_labels, title, headers=ADHAWK_MAIN_SHEET_HEADERS, col_widths=ADHAWK_MAIN_SHEET_WIDTHS)

    parties = sorted(tree.items(), key=lambda kv: -sum(_sum_lists(
        [s for adv in kv[1].values() for tp in adv.values() for s in tp.values()], n_weeks
    )))

    for party, advertisers in parties:
        w.set_party(party)
        adv_totals = {
            adv: _sum_lists([s for tp in types.values() for s in tp.values()], n_weeks)
            for adv, types in advertisers.items()
        }
        ranked_advs = sorted(advertisers.items(), key=lambda kv: -sum(adv_totals[kv[0]]))

        party_weekly = _sum_lists(adv_totals.values(), n_weeks)

        for adv, types in ranked_advs:
            adv_start = w.row
            w.write_group_label(1, adv_start, adv)

            ranked_types = sorted(
                types.items(), key=lambda kv: -sum(sum(weekly) for weekly in kv[1].values()),
            )

            for mtype, stations in ranked_types:
                type_start = w.row
                w.write_group_label(2, type_start, mtype)
                type_weekly = _sum_lists(stations.values(), n_weeks)

                if mtype.strip().upper() == "CTV":
                    # Same collapsing as the AdImpact sheet: CTV never
                    # breaks out by platform, it's just one line.
                    w.write_leaf(mtype, sum(type_weekly), type_weekly)
                else:
                    ranked_stations = sorted(stations.items(), key=lambda kv: -sum(kv[1]))
                    for station, weekly in ranked_stations:
                        w.write_leaf(station, sum(weekly), weekly)
                    w.write_subtotal(f"{mtype} Total", sum(type_weekly), type_weekly)
                type_end = w.row - 1
                w.merge_col(2, type_start, type_end)  # TYPE column

            w.write_advertiser_total(f"{adv} Total", sum(adv_totals[adv]), adv_totals[adv])
            adv_end = w.row - 1
            w.merge_col(1, adv_start, adv_end)  # CANDIDATE / COMMITTEE column
            w.blank_row()

        w.write_party_total(f"{party.upper()} PARTY TOTAL", sum(party_weekly), party_weekly)
        w.merge_label_row(w.row - 1)

    grand_weekly = _sum_lists(
        [_sum_lists(
            [s for adv in advertisers.values() for tp in adv.values() for s in tp.values()],
            n_weeks,
        ) for _, advertisers in parties],
        n_weeks,
    )
    w.write_grand_total("GRAND TOTAL", sum(grand_weekly), grand_weekly)
    w.merge_label_row(w.row - 1)
    w.write_footer()
    return ws


def _write_market_summary_sheet(wb, tree, n_weeks):
    ws = wb.create_sheet("Market Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16

    c = ws.cell(1, 1, "Market Summary")
    c.font = Font(name=FONT_NAME, bold=True, size=14, color="FFFFFF")
    c.fill = _fill(NAVY)
    ws.merge_cells("A1:B1")

    parties = sorted(tree.items(), key=lambda kv: -sum(_sum_lists(
        [s for adv in kv[1].values() for mk in adv.values() for tp in mk.values() for s in tp.values()], n_weeks
    )))

    row = 3
    for party, advertisers in parties:
        ramp = _ramp_for(party)
        adv_totals = {
            adv: sum(_sum_lists([s for mk in markets.values() for tp in mk.values() for s in tp.values()], n_weeks))
            for adv, markets in advertisers.items()
        }
        ranked_advs = sorted(advertisers.items(), key=lambda kv: -adv_totals[kv[0]])

        for adv, markets in ranked_advs:
            row += 1
            c = ws.cell(row, 1, adv)
            c.font = Font(name=FONT_NAME, bold=True, size=11, color=TEXT_DARK)
            row += 1
            for col, text in enumerate(["MARKET", "SPEND"], start=1):
                c = ws.cell(row, col, text)
                c.font = Font(name=FONT_NAME, bold=True, size=9, color="FFFFFF")
                c.fill = _fill(NAVY)
            row += 1

            # Group by media type first (all CTV markets together, all
            # Digital markets together), each with its own type total,
            # biggest type first. Keep each market's per-platform stations
            # around (not just its total) so Digital can break Facebook,
            # Google, etc. out the same way the main sheet does — CTV stays
            # collapsed to one line per market, matching the main sheet.
            by_type = {}
            for market, types in markets.items():
                for mtype, stations in types.items():
                    by_type.setdefault(mtype, {})[market] = stations
            ranked_types = sorted(
                by_type.items(),
                key=lambda kv: -sum(sum(_sum_lists(stations.values(), n_weeks)) for stations in kv[1].values()),
            )

            for mtype, market_map in ranked_types:
                c = ws.cell(row, 1, mtype.upper())
                c.font = Font(name=FONT_NAME, bold=True, size=9, color=TEXT_DARK)
                row += 1
                is_ctv = mtype.strip().upper() == "CTV"
                market_entries = [
                    (market, sum(_sum_lists(stations.values(), n_weeks)), stations)
                    for market, stations in market_map.items()
                ]
                for market, total, stations in sorted(market_entries, key=lambda x: -x[1]):
                    if is_ctv:
                        ws.cell(row, 1, market).font = Font(name=FONT_NAME, size=9, color=TEXT_DARK)
                        e = ws.cell(row, 2, total)
                        e.number_format = TOTAL_CURRENCY
                        e.font = Font(name=FONT_NAME, size=9, color=TEXT_DARK)
                        row += 1
                    else:
                        # Digital: surface the platform breakdown (Facebook,
                        # Google, ...) under each market, same as the main
                        # sheet, instead of collapsing straight to a total.
                        ranked_platforms = sorted(stations.items(), key=lambda kv: -sum(kv[1]))
                        for platform, weekly in ranked_platforms:
                            pc = ws.cell(row, 1, platform)
                            pc.font = Font(name=FONT_NAME, size=9, color=TEXT_DARK)
                            pc.alignment = Alignment(horizontal="left", indent=1)
                            pc.fill = _fill(ramp["leaf"])
                            e = ws.cell(row, 2, sum(weekly))
                            e.number_format = TOTAL_CURRENCY
                            e.font = Font(name=FONT_NAME, size=9, color=TEXT_DARK)
                            e.fill = _fill(ramp["leaf"])
                            row += 1
                        c = ws.cell(row, 1, f"{market} Total")
                        c.font = Font(name=FONT_NAME, bold=True, size=9, color=TEXT_DARK)
                        c.fill = _fill(ramp["subtotal"])
                        e = ws.cell(row, 2, total)
                        e.number_format = TOTAL_CURRENCY
                        e.font = Font(name=FONT_NAME, bold=True, size=9, color=TEXT_DARK)
                        e.fill = _fill(ramp["subtotal"])
                        row += 1
                type_total = sum(t for _, t, _ in market_entries)
                c = ws.cell(row, 1, f"{mtype} Total")
                c.font = Font(name=FONT_NAME, bold=True, size=9, color=TEXT_DARK)
                c.fill = _fill(ramp["type_total"])
                e = ws.cell(row, 2, type_total)
                e.number_format = TOTAL_CURRENCY
                e.font = Font(name=FONT_NAME, bold=True, size=9, color=TEXT_DARK)
                e.fill = _fill(ramp["type_total"])
                row += 1

            c = ws.cell(row, 1, f"{adv} — Total")
            c.font = Font(name=FONT_NAME, bold=True, size=9, color=TEXT_DARK)
            c.fill = _fill(ramp["adv_total"])
            e = ws.cell(row, 2, adv_totals[adv])
            e.number_format = TOTAL_CURRENCY
            e.font = Font(name=FONT_NAME, bold=True, size=9, color=TEXT_DARK)
            e.fill = _fill(ramp["adv_total"])
            row += 2
    return ws


CHANGE_CURRENCY = '+$#,##0;-$#,##0;$0'
CHANGE_PERCENT = "+0.0%;-0.0%;0.0%"

THIS_WEEK_HEADERS = [
    "CANDIDATE / COMMITTEE", "MARKET", "TYPE", "STATION / PLATFORM",
    "THIS WEEK", "PRIOR WEEK", "CHANGE ($)", "CHANGE (%)",
]
THIS_WEEK_HEADERS_ADHAWK = [
    "CANDIDATE / COMMITTEE", "TYPE", "PLATFORM",
    "THIS WEEK", "PRIOR WEEK", "CHANGE ($)", "CHANGE (%)",
]


def _write_week_row(ws, row, label_col, label, this_val, prior_val, fill, bold, text_color, border, font_size=9,
                     value_col=5, last_col=8):
    font = Font(name=FONT_NAME, bold=bold, size=font_size, color=text_color)
    for col in range(1, last_col + 1):
        c = ws.cell(row, col)
        c.fill = _fill(fill)
        c.border = border
        c.font = font
    if label is not None:
        ws.cell(row, label_col, label)
    ws.cell(row, value_col, this_val).number_format = TOTAL_CURRENCY
    change = this_val - prior_val
    ws.cell(row, value_col + 1, prior_val).number_format = TOTAL_CURRENCY
    ws.cell(row, value_col + 2, change).number_format = CHANGE_CURRENCY
    if prior_val > 0:
        ws.cell(row, value_col + 3, change / prior_val).number_format = CHANGE_PERCENT
    elif this_val > 0:
        ws.cell(row, value_col + 3, "New")
    return row + 1


def _write_this_week_sheet(wb, tree, week_iso, this_idx):
    ws = wb.create_sheet("This Week")
    ws.sheet_view.showGridLines = False

    prior_idx = this_idx - 1 if this_idx > 0 else None
    week_start = datetime.strptime(week_iso[this_idx], "%Y-%m-%d")
    week_end = week_start + timedelta(days=6)
    week_range = f"{week_start.strftime('%B %d')} – {week_end.strftime('%B %d, %Y')}"

    all_this, all_prior = _sum_pairs(
        (weekly[this_idx], weekly[prior_idx] if prior_idx is not None else 0.0)
        for advertisers in tree.values()
        for markets in advertisers.values()
        for types in markets.values()
        for stations in types.values()
        for weekly in stations.values()
    )
    total_advertisers = sum(len(advertisers) for advertisers in tree.values())

    filtered = _filter_this_week(tree, this_idx, prior_idx)
    active_advertisers = sum(len(advertisers) for advertisers in filtered.values())

    last_col = 8
    for col in range(1, last_col + 1):
        ws.cell(1, col).fill = _fill(NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(1, 1, f"THIS WEEK — {week_range}")
    c.font = Font(name=FONT_NAME, bold=True, size=16, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 49.5

    change = all_this - all_prior
    pct = (change / all_prior * 100) if all_prior else None
    arrow = "↑" if change >= 0 else "↓"
    pct_txt = f"{arrow} {abs(pct):.0f}% vs prior week" if pct is not None else "no prior-week spend to compare"
    summary = (
        f"${all_this:,.0f} total spend this week  |  {pct_txt} (${all_prior:,.0f})  |  "
        f"{active_advertisers} of {total_advertisers} advertisers active"
    )
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    c = ws.cell(3, 1, summary)
    c.font = Font(name=FONT_NAME, bold=True, size=11, color=TEXT_DARK)

    header_row = 5
    for col, text in enumerate(THIS_WEEK_HEADERS, start=1):
        c = ws.cell(header_row, col, text)
        c.font = Font(name=FONT_NAME, bold=True, size=9, color="FFFFFF")
        c.fill = _fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _medium_border(HEADER_ACCENT)

    widths = {"A": 28, "B": 24, "C": 10, "D": 30, "E": 14, "F": 14, "G": 14, "H": 12}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = f"A{header_row + 1}"

    row = header_row + 1
    parties = sorted(filtered.items(), key=lambda kv: -_sum_pairs(
        (s[0], s[1]) for adv in kv[1].values() for mk in adv.values() for tp in mk.values() for s in tp.values()
    )[0])

    for party, advertisers in parties:
        ramp = _ramp_for(party)
        adv_totals = {
            adv: _sum_pairs((s[0], s[1]) for mk in markets.values() for tp in mk.values() for s in tp.values())
            for adv, markets in advertisers.items()
        }
        ranked_advs = sorted(advertisers.items(), key=lambda kv: -adv_totals[kv[0]][0])
        party_total = _sum_pairs(adv_totals.values())

        if not ranked_advs:
            continue

        for adv, markets in ranked_advs:
            adv_start = row
            ws.cell(adv_start, 1, adv).font = Font(name=FONT_NAME, bold=True, size=9, color=TEXT_DARK)
            ws.cell(adv_start, 1).alignment = Alignment(vertical="center", wrap_text=True)

            by_type = {}
            for market, types in markets.items():
                for mtype, stations in types.items():
                    by_type.setdefault(mtype, {})[market] = stations
            ranked_types = sorted(
                by_type.items(),
                key=lambda kv: -_sum_pairs(s for st in kv[1].values() for s in st.values())[0],
            )

            for mtype, market_map in ranked_types:
                for market, stations in market_map.items():
                    market_start = row
                    ws.cell(market_start, 2, market).font = Font(name=FONT_NAME, bold=True, size=9, color=TEXT_DARK)
                    ws.cell(market_start, 2).alignment = Alignment(vertical="center", wrap_text=True)
                    type_start = row
                    ws.cell(type_start, 3, mtype).font = Font(name=FONT_NAME, bold=True, size=9, color=TEXT_DARK)
                    ws.cell(type_start, 3).alignment = Alignment(horizontal="center", vertical="center")
                    type_total = _sum_pairs(stations.values())

                    if mtype.strip().upper() == "CTV":
                        row = _write_week_row(
                            ws, row, 4, f"{market} - {mtype}", *type_total,
                            ramp["leaf"], False, TEXT_DARK, _thin_border(BORDER_LIGHT),
                        )
                    else:
                        ranked_stations = sorted(stations.items(), key=lambda kv: -kv[1][0])
                        for station, (tv, pv) in ranked_stations:
                            row = _write_week_row(
                                ws, row, 4, station, tv, pv,
                                ramp["leaf"], False, TEXT_DARK, _thin_border(BORDER_LIGHT),
                            )
                        row = _write_week_row(
                            ws, row, 4, f"{market} - {mtype} Total", *type_total,
                            ramp["subtotal"], True, TEXT_DARK, _thin_border(BORDER_MED),
                        )
                    if row - 1 > market_start:
                        ws.merge_cells(start_row=market_start, start_column=2, end_row=row - 1, end_column=2)
                    if mtype.strip().upper() != "CTV" and row - 1 > type_start:
                        ws.merge_cells(start_row=type_start, start_column=3, end_row=row - 1, end_column=3)

                type_grand = _sum_pairs(s for st in market_map.values() for s in st.values())
                row = _write_week_row(
                    ws, row, 4, f"{mtype.upper()} TOTAL", *type_grand,
                    ramp["type_total"], True, TEXT_DARK, _thin_border(BORDER_MED),
                )

            row = _write_week_row(
                ws, row, 4, f"{adv} Total", *adv_totals[adv],
                ramp["adv_total"], True, TEXT_DARK, _medium_border(BORDER_MED),
            )
            if row - 1 > adv_start:
                ws.merge_cells(start_row=adv_start, start_column=1, end_row=row - 1, end_column=1)
            row += 1  # blank spacer

        row = _write_week_row(
            ws, row, 1, f"{party.upper()} PARTY TOTAL", *party_total,
            ramp["party_total"], True, "FFFFFF", _medium_border("FFFFFF"), font_size=10,
        )
        ws.merge_cells(start_row=row - 1, start_column=1, end_row=row - 1, end_column=4)

    row = _write_week_row(
        ws, row, 1, "GRAND TOTAL", all_this, all_prior,
        GRAND_FILL, True, "FFFFFF", _medium_border("FFFFFF"), font_size=10,
    )
    ws.merge_cells(start_row=row - 1, start_column=1, end_row=row - 1, end_column=4)

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row, 1, "Report prepared by GPS Impact  |  Confidential")
    c.font = Font(name=FONT_NAME, italic=True, size=8, color=FOOTER_GRAY)
    ws.row_dimensions[row].height = 18
    return ws


def _write_this_week_sheet_adhawk(wb, tree, week_iso, this_idx):
    """`_write_this_week_sheet`, minus the Market column."""
    ws = wb.create_sheet("This Week")
    ws.sheet_view.showGridLines = False

    prior_idx = this_idx - 1 if this_idx > 0 else None
    week_start = datetime.strptime(week_iso[this_idx], "%Y-%m-%d")
    week_end = week_start + timedelta(days=6)
    week_range = f"{week_start.strftime('%B %d')} – {week_end.strftime('%B %d, %Y')}"

    all_this, all_prior = _sum_pairs(
        (weekly[this_idx], weekly[prior_idx] if prior_idx is not None else 0.0)
        for advertisers in tree.values()
        for types in advertisers.values()
        for stations in types.values()
        for weekly in stations.values()
    )
    total_advertisers = sum(len(advertisers) for advertisers in tree.values())

    filtered = _filter_this_week_no_market(tree, this_idx, prior_idx)
    active_advertisers = sum(len(advertisers) for advertisers in filtered.values())

    last_col = 7
    for col in range(1, last_col + 1):
        ws.cell(1, col).fill = _fill(NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(1, 1, f"THIS WEEK — {week_range}")
    c.font = Font(name=FONT_NAME, bold=True, size=16, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 49.5

    change = all_this - all_prior
    pct = (change / all_prior * 100) if all_prior else None
    arrow = "↑" if change >= 0 else "↓"
    pct_txt = f"{arrow} {abs(pct):.0f}% vs prior week" if pct is not None else "no prior-week spend to compare"
    summary = (
        f"${all_this:,.0f} total spend this week  |  {pct_txt} (${all_prior:,.0f})  |  "
        f"{active_advertisers} of {total_advertisers} advertisers active"
    )
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    c = ws.cell(3, 1, summary)
    c.font = Font(name=FONT_NAME, bold=True, size=11, color=TEXT_DARK)

    header_row = 5
    for col, text in enumerate(THIS_WEEK_HEADERS_ADHAWK, start=1):
        c = ws.cell(header_row, col, text)
        c.font = Font(name=FONT_NAME, bold=True, size=9, color="FFFFFF")
        c.fill = _fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _medium_border(HEADER_ACCENT)

    widths = {"A": 32, "B": 10, "C": 24, "D": 14, "E": 14, "F": 14, "G": 12}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = f"A{header_row + 1}"

    row = header_row + 1
    parties = sorted(filtered.items(), key=lambda kv: -_sum_pairs(
        (s[0], s[1]) for adv in kv[1].values() for tp in adv.values() for s in tp.values()
    )[0])

    for party, advertisers in parties:
        ramp = _ramp_for(party)
        adv_totals = {
            adv: _sum_pairs((s[0], s[1]) for tp in types.values() for s in tp.values())
            for adv, types in advertisers.items()
        }
        ranked_advs = sorted(advertisers.items(), key=lambda kv: -adv_totals[kv[0]][0])
        party_total = _sum_pairs(adv_totals.values())

        if not ranked_advs:
            continue

        for adv, types in ranked_advs:
            adv_start = row
            ws.cell(adv_start, 1, adv).font = Font(name=FONT_NAME, bold=True, size=9, color=TEXT_DARK)
            ws.cell(adv_start, 1).alignment = Alignment(vertical="center", wrap_text=True)

            ranked_types = sorted(
                types.items(),
                key=lambda kv: -_sum_pairs(kv[1].values())[0],
            )

            for mtype, stations in ranked_types:
                type_start = row
                ws.cell(type_start, 2, mtype).font = Font(name=FONT_NAME, bold=True, size=9, color=TEXT_DARK)
                ws.cell(type_start, 2).alignment = Alignment(horizontal="center", vertical="center")
                type_total = _sum_pairs(stations.values())

                if mtype.strip().upper() == "CTV":
                    row = _write_week_row(
                        ws, row, 3, mtype, *type_total,
                        ramp["leaf"], False, TEXT_DARK, _thin_border(BORDER_LIGHT),
                        value_col=4, last_col=7,
                    )
                else:
                    ranked_stations = sorted(stations.items(), key=lambda kv: -kv[1][0])
                    for station, (tv, pv) in ranked_stations:
                        row = _write_week_row(
                            ws, row, 3, station, tv, pv,
                            ramp["leaf"], False, TEXT_DARK, _thin_border(BORDER_LIGHT),
                            value_col=4, last_col=7,
                        )
                    row = _write_week_row(
                        ws, row, 3, f"{mtype} Total", *type_total,
                        ramp["subtotal"], True, TEXT_DARK, _thin_border(BORDER_MED),
                        value_col=4, last_col=7,
                    )
                if row - 1 > type_start:
                    ws.merge_cells(start_row=type_start, start_column=2, end_row=row - 1, end_column=2)

            row = _write_week_row(
                ws, row, 3, f"{adv} Total", *adv_totals[adv],
                ramp["adv_total"], True, TEXT_DARK, _medium_border(BORDER_MED),
                value_col=4, last_col=7,
            )
            if row - 1 > adv_start:
                ws.merge_cells(start_row=adv_start, start_column=1, end_row=row - 1, end_column=1)
            row += 1  # blank spacer

        row = _write_week_row(
            ws, row, 1, f"{party.upper()} PARTY TOTAL", *party_total,
            ramp["party_total"], True, "FFFFFF", _medium_border("FFFFFF"), font_size=10,
            value_col=4, last_col=7,
        )
        ws.merge_cells(start_row=row - 1, start_column=1, end_row=row - 1, end_column=3)

    row = _write_week_row(
        ws, row, 1, "GRAND TOTAL", all_this, all_prior,
        GRAND_FILL, True, "FFFFFF", _medium_border("FFFFFF"), font_size=10,
        value_col=4, last_col=7,
    )
    ws.merge_cells(start_row=row - 1, start_column=1, end_row=row - 1, end_column=3)

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row, 1, "Report prepared by GPS Impact  |  Confidential")
    c.font = Font(name=FONT_NAME, italic=True, size=8, color=FOOTER_GRAY)
    ws.row_dimensions[row].height = 18
    return ws


def _write_no_data_week_sheet(wb, target_iso):
    ws = wb.create_sheet("This Week")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 60
    ws.row_dimensions[1].height = 40

    week_start = datetime.strptime(target_iso, "%Y-%m-%d")
    week_end = week_start + timedelta(days=6)
    week_range = f"{week_start.strftime('%B %d')} – {week_end.strftime('%B %d, %Y')}"

    ws.merge_cells("A1:C1")
    c = ws.cell(1, 1, f"THIS WEEK — {week_range}")
    c.font = Font(name=FONT_NAME, bold=True, size=14, color="FFFFFF")
    c.fill = _fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in (2, 3):
        ws.cell(1, col).fill = _fill(NAVY)

    c2 = ws.cell(2, 1, "No spending data available for this week yet.")
    c2.font = Font(name=FONT_NAME, italic=True, size=10, color=FOOTER_GRAY)
    return ws


CREATIVE_HEADERS = ["CANDIDATE / COMMITTEE", "CREATIVE", "PLATFORM", "TONE", "TOTAL SPEND", "START DATE", "END DATE"]
CREATIVE_DATE_FORMAT = "m/d"


def _write_creative_timeline_sheet(wb, creative_rows, party_lookup, week_iso, week_labels):
    """One row per creative, grouped by party then advertiser, with a
    shaded bar across every media week the creative's flight overlapped.
    A shaded week means "live at some point that week," not necessarily
    the full week — flights rarely start or end on a Tuesday.
    """
    ws = wb.create_sheet("Creative Timeline")
    ws.sheet_view.showGridLines = False

    n_weeks = len(week_iso)
    n_fixed = len(CREATIVE_HEADERS)
    last_col = n_fixed + n_weeks
    frozen_cols = n_fixed

    for col in range(1, last_col + 1):
        ws.cell(1, col).fill = _fill(NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=frozen_cols)
    c = ws.cell(1, 1, "CREATIVE TIMELINE")
    c.font = Font(name=FONT_NAME, bold=True, size=14, color="FFFFFF")
    c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True, indent=1)
    if last_col > frozen_cols:
        ws.merge_cells(start_row=1, start_column=frozen_cols + 1, end_row=1, end_column=last_col)
    ws.row_dimensions[1].height = 49.5
    if LOGO_FILE.exists():
        logo = XLImage(str(LOGO_FILE))
        aspect = logo.width / logo.height
        logo.height = 34
        logo.width = 34 * aspect
        ws.add_image(logo, "A1")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    c = ws.cell(2, 1, "A shaded week means the creative was live at some point during that week, not necessarily the full week.")
    c.font = Font(name=FONT_NAME, italic=True, size=9, color=FOOTER_GRAY)

    header_row = 4
    for col, text in enumerate(CREATIVE_HEADERS + week_labels, start=1):
        c = ws.cell(header_row, col, text)
        c.font = Font(name=FONT_NAME, bold=True, size=9, color="FFFFFF")
        c.fill = _fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _medium_border(HEADER_ACCENT)
    ws.row_dimensions[header_row].height = 31.5

    widths = {"A": 28, "B": 30, "C": 10, "D": 12, "E": 14, "F": 7, "G": 7}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w
    for i in range(n_weeks):
        ws.column_dimensions[get_column_letter(n_fixed + 1 + i)].width = 5.5
    ws.freeze_panes = f"{get_column_letter(frozen_cols + 1)}{header_row + 1}"

    by_party = {}
    for cr in creative_rows:
        party = party_lookup.get(cr["advertiser"])
        by_party.setdefault(party, {}).setdefault(cr["advertiser"], []).append(cr)
    ranked_parties = sorted(
        by_party.items(), key=lambda kv: -sum(cr["total_spend"] for advs in kv[1].values() for cr in advs)
    )

    row = header_row + 1
    for party, advertisers in ranked_parties:
        ramp = _ramp_for(party)
        ranked_advs = sorted(
            advertisers.items(), key=lambda kv: -sum(cr["total_spend"] for cr in kv[1])
        )
        for adv, creatives in ranked_advs:
            adv_start = row
            creatives = sorted(creatives, key=lambda cr: cr["start"])
            for cr in creatives:
                font = Font(name=FONT_NAME, size=9, color=TEXT_DARK)
                ws.cell(row, 1, adv).font = font
                title_cell = ws.cell(row, 2, cr["title"])
                if cr.get("url"):
                    title_cell.hyperlink = cr["url"]
                    title_cell.font = Font(name=FONT_NAME, size=9, color=BRAND_BLUE, underline="single")
                else:
                    title_cell.font = font
                ws.cell(row, 3, cr["platform"]).font = font
                ws.cell(row, 3).alignment = Alignment(horizontal="center")
                ws.cell(row, 4, cr["tone"] or "").font = font
                e = ws.cell(row, 5, cr["total_spend"])
                e.number_format = TOTAL_CURRENCY
                e.font = font
                for col, dt in ((6, cr["start"]), (7, cr["end"])):
                    dc = ws.cell(row, col, dt)
                    dc.number_format = CREATIVE_DATE_FORMAT
                    dc.font = font
                for col in range(1, n_fixed + 1):
                    ws.cell(row, col).border = _thin_border(BORDER_LIGHT)
                bar_fill = _fill(ramp["adv_total"])
                for i in parse.creative_active_weeks(cr, week_iso):
                    bc = ws.cell(row, n_fixed + 1 + i)
                    bc.fill = bar_fill
                    bc.border = _thin_border(BORDER_LIGHT)
                row += 1
            if row - 1 > adv_start:
                ws.merge_cells(start_row=adv_start, start_column=1, end_row=row - 1, end_column=1)
            row += 1  # blank spacer

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row, 1, "Report prepared by GPS Impact  |  Confidential")
    c.font = Font(name=FONT_NAME, italic=True, size=8, color=FOOTER_GRAY)
    ws.row_dimensions[row].height = 18
    return ws


def write_excel_report(leaf_rows, index_map, week_labels, output_path, title="DIGITAL COMPETITIVE REPORT",
                        week_iso=None, this_week_iso=None, creative_rows=None):
    """`this_week_iso` is the target Tuesday (YYYY-MM-DD) to show on the
    "This Week" tab — pass `parse.current_media_week_iso()` for "today's"
    media week, or a specific date to pin a historical week. If that date
    isn't present in the export, a placeholder tab is shown instead of
    silently substituting some other week.

    `creative_rows`, if given (from `parse.load_creative_export`), adds a
    Creative Timeline tab. Party coloring for it comes from `leaf_rows` —
    the Topline Creatives export has no Party column of its own.
    """
    n_weeks = len(week_labels)
    tree = build_hierarchy(leaf_rows, index_map, n_weeks)
    wb = Workbook()
    _write_main_sheet(wb, tree, week_labels, n_weeks, title)
    _write_market_summary_sheet(wb, tree, n_weeks)
    if this_week_iso is not None and week_iso is not None:
        this_idx = week_iso.index(this_week_iso) if this_week_iso in week_iso else None
        if this_idx is not None:
            _write_this_week_sheet(wb, tree, week_iso, this_idx)
        else:
            _write_no_data_week_sheet(wb, this_week_iso)
    if creative_rows:
        party_lookup = {r["advertiser"]: r["party"] for r in leaf_rows}
        creative_week_iso, creative_week_labels = parse.build_creative_week_axis(creative_rows)
        _write_creative_timeline_sheet(wb, creative_rows, party_lookup, creative_week_iso, creative_week_labels)
    wb.active = 0
    wb.save(output_path)


def write_adhawk_report(leaf_rows, week_labels, output_path, title="DIGITAL COMPETITIVE REPORT",
                         week_iso=None, this_week_iso=None):
    """Same idea as `write_excel_report`, for AdHawk-sourced data: no Market
    Summary tab (AdHawk has no market dimension to summarize by), and the
    main/This Week sheets both drop the Market column entirely.
    """
    n_weeks = len(week_labels)
    tree = build_hierarchy_no_market(leaf_rows, n_weeks)
    wb = Workbook()
    _write_main_sheet_adhawk(wb, tree, week_labels, n_weeks, title)
    if this_week_iso is not None and week_iso is not None:
        this_idx = week_iso.index(this_week_iso) if this_week_iso in week_iso else None
        if this_idx is not None:
            _write_this_week_sheet_adhawk(wb, tree, week_iso, this_idx)
        else:
            _write_no_data_week_sheet(wb, this_week_iso)
    wb.active = 0
    wb.save(output_path)
